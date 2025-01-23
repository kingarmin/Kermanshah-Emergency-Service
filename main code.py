import tkinter as tk
from tkinter import filedialog
import pandas as pd
from openpyxl import load_workbook
from pathlib import Path
import os
Input_path=None
Output_path=None
last_output_path=None
def read_excel_sheet_with_column_names(file_path, sheet_name):
  try:
    workbook = load_workbook(file_path,data_only=True)
    sheet = workbook[sheet_name]
    column_names = [cell.value for cell in sheet[3]] 
    data = []
    for row in sheet.iter_rows(min_row=4, values_only=True): 
      data.append(row)
    df = pd.DataFrame(data, columns=column_names) 
    return df
  except KeyError:
    print(f"Sheet '{sheet_name}' not found in the Excel file.")
    return None
  
def select_path(x):
  global Output_path
  global Input_path

  if x == 'Input':
    path = filedialog.askopenfilename(title="Select Input File")
    if path:
      Input_path = path
      Input_file_label.config(text=f"Selected Input file address: {path}")
  elif x == 'Output':
    path = filedialog.askdirectory(title="Select Output Directory")
    if path:
      Output_path = path
      Output_file_label.config(text=f"Selected Output folder address: {path}")
def process():
  global Input_path
  global Output_path
  global last_output_path
  if Input_path==None or Output_path==None:
    Eror_status_label.config(text="you didn't selecht Input file address or Output file address")
  else:
    sheet1 = read_excel_sheet_with_column_names(Input_path, 'ارزیابی')
    sheet2 = read_excel_sheet_with_column_names(Input_path, 'کارکنان')
    sheet1.drop(columns=[None],inplace=True)
    sheet2.drop(columns=[None],inplace=True)
    sheet1=sheet1[['کارکرد دوره (ساعت)','تعداد ماموریت شهری','تعداد ماموریت جاده ای','تعداد تماس در دوره','تعداد بازدید میدانی در دوره','تعداد فرماندهی حادثه در دوره','رضایت']]
    sheet2=sheet2[['شماره پرسنلی','نام','نام خانوادگی','شغل','رتبه','امتیاز رتبه','جایگاه سازمانی','مدرک تحصیلی','منطقه خدمتی','سابقه (سال)']]
    df=pd.concat([sheet2,sheet1.loc[list(range(0,len(list(sheet2.index))-1))]],axis=1)
    def rezaiat(x):
      if x>=0 and x<=50:
        return 0.6
      if x>50 and x<=60:
        return 1
      if x>60 and x<=70:
        return 1.6
      if x>70 and x<=80:
        return 2.5
      if x>80 and x<=90:
        return 3.6
      else:
        return 5
    x = df.copy()
    mapping = {'عملیاتی': 3, 'دیسپچ': 2, 'ستاد': 1}
    x['جایگاه سازمانی'] = x['جایگاه سازمانی'].map(mapping)
    mapping={'ارشد مرتبط':3 , 'کارشناسی مرتبط': 2.5 , 'کاردانی': 2 , 'دیپلم':1.5 , 'زیر دیپلم': 0.5}
    x['مدرک تحصیلی']=x['مدرک تحصیلی'].map(mapping)
    x['سابقه (سال)']=x['سابقه (سال)'].astype(float)*0.25
    x['کارکرد دوره (ساعت)']=x['کارکرد دوره (ساعت)'].astype(float)*(0.1/8)
    mapping={'شهری':1,'جاده ای برخوردار':2 , 'جاده ای غیر برخوردار':4}
    x['منطقه خدمتی']=x['منطقه خدمتی'].map(mapping)
    x['رضایت']=x['رضایت'].transform(rezaiat)
    x['عملکرد ماهیانه']=x['تعداد ماموریت جاده ای']*(0.1/3)+x['تعداد ماموریت شهری']*(0.1/7)+x['تعداد بازدید میدانی در دوره']*1+x['تعداد تماس در دوره']*1
    df['امتیاز نهایی']=(x['جایگاه سازمانی']+x['مدرک تحصیلی']+x['سابقه (سال)']+x['کارکرد دوره (ساعت)']+x['منطقه خدمتی']+x['رضایت'])*df['امتیاز رتبه']*x['عملکرد ماهیانه']
    sheet3 = read_excel_sheet_with_column_names(Input_path, 'تنظیمات دوره')
    sheet3.index=sheet3['تعداد روز ماه'].to_list()
    sheet3.drop(columns=['تعداد روز ماه'],inplace=True)
    sheet3=sheet3[30].loc['مبلغ بودجه']
    x['امتیاز نهایی']=df['امتیاز نهایی']
    y=x[x['جایگاه سازمانی']==2]
    x['2']=y['امتیاز نهایی']
    y=x[x['جایگاه سازمانی']==3]
    x['3']=y['امتیاز نهایی']
    bodge=(x['3'].sum()+x['2'].sum())/sheet3
    x.drop(columns=['2','3'],inplace=True)
    df['مبلغ کارانه']=df['امتیاز نهایی']*bodge
    if os.path.exists(Output_path+'/Output.xlsx'):
      os.remove(Output_path+'/Output.xlsx')
    df.to_excel(Output_path+'/Output.xlsx', index=False)
    last_output_path=Output_path
    Output_path=None
    Input_path=None
    Eror_status_label.config(text='')

def open_file():
  if last_output_path!=None:
    os.startfile(last_output_path+'/Output.xlsx')
root = tk.Tk()
root.title("Kermanshah Emergency")
input_button = tk.Button(root, text="Input file address", command=lambda : select_path('Input'))
input_button.pack(pady=20,padx=200)
output_button = tk.Button(root, text="Output file address", command=lambda : select_path('Output'))
output_button.pack(pady=20,padx=200)

Input_file_label = tk.Label(root, text="Input file address :No file selected")
Input_file_label.pack(pady=20)
Output_file_label = tk.Label(root, text="Output file address :No file selected")
Output_file_label.pack(pady=20)


start_button = tk.Button(root, text="start_process", command=process)
start_button.pack(pady=20,padx=200)
open_output_button = tk.Button(root, text="Open output file", command=open_file)
open_output_button.pack(pady=20,padx=200)
Eror_status_label = tk.Label(root, text="")
Eror_status_label.pack(pady=20)

root.mainloop()
